from openpyxl import Workbook
from openpyxl import load_workbook

# adjust classNum type (int/str) 
CLASS_NUM_COL = 3
TOTAL_SCORE_COL = 5
EXAM = "高一上期末"
# at most 6 classes per part, check var a in function write()
PARTS = ((7,8,9,10,11),(12,13,14,15),(16,17))
LEVEL = (6000,11000,19000,21000)

def level(classNum):
    for i in PARTS:
        if classNum in i:
            a = PARTS.index(i)
            return LEVEL[a], LEVEL[a+1]

# Counts the students whose rank of the subject is higher than his total score rank
def counts(rankList, levels):
    a, b = 0, 0
    for i in rankList:
        if i[0] <= levels[0] and i[1] <= levels[0]:
                a += 1
        elif i[0] <= levels[1] and i[1] <= levels[1]:
                b += 1
    return a, b

# Generate a list of students' total and subject ranks, remove absent students'
def generate(classNum,subjectNum):
    rankList = []
    for i in ranks.values:
        if i[CLASS_NUM_COL] == classNum:
            rankList.append([i[TOTAL_SCORE_COL], i[subjectNum]])
    rankList = [i for i in rankList if i[0] != "" and i[1] != ""]
    return rankList

def write(numList):
    a = 13 * PARTS.index(numList) + 2
    results.active.append(["班级","分段","语文","数学","英语","物理","化学","生物","政治","历史","地理"])
    for i in numList:
        results.active.cell(row=a+i-numList[0], column=1, value=i)
        results.active.cell(row=a+i-numList[0]+len(numList), column=1, value=i)
        results.active.cell(row=a, column=2, value=level(i)[0])
        results.active.cell(row=a+len(numList), column=2, value=level(i)[1])
        for j in range(9):
            rankList = generate(i,TOTAL_SCORE_COL+4*(1+j))
            result = counts(rankList, level(i))
            results.active.cell(row=a+i-numList[0], column=3+j, value=result[0])
            results.active.cell(row=a+i-numList[0]+len(numList), column=3+j, value=result[1])

book = load_workbook(f"成绩单/{EXAM}.xlsx")
ranks = book.active
results = Workbook()
for i in PARTS:
    write(i)
results.save(f"{EXAM}基数.xlsx")